# file_operations.py
"""File upload, export, and preset management."""

import io
import json
import re
from datetime import datetime, timezone
from typing import Dict, Optional, Tuple

import pandas as pd
import streamlit as st


@st.cache_data(show_spinner=False)
def read_csv_file(file) -> pd.DataFrame:
    """Read CSV file with fallback encoding.

    Cached on the uploaded file's contents so re-running the Streamlit
    script (every widget change) doesn't re-parse the same CSV.

    Args:
        file: Uploaded file object

    Returns:
        DataFrame
    """
    try:
        return pd.read_csv(file)
    except (UnicodeDecodeError, pd.errors.ParserError):
        file.seek(0)
        return pd.read_csv(file, encoding="utf-8", engine="python")


@st.cache_data(show_spinner=False)
def read_excel_file(file, sheet_name=None) -> Dict[str, pd.DataFrame]:
    """Read Excel file with fallback.

    Cached on the uploaded file's contents; see :func:`read_csv_file`.

    Args:
        file: Uploaded file object
        sheet_name: Specific sheet name or None for all sheets

    Returns:
        Dict of sheet_name -> DataFrame
    """
    try:
        return pd.read_excel(file, sheet_name=sheet_name)
    except (ValueError, OSError):
        file.seek(0)
        return pd.read_excel(file, sheet_name=sheet_name)


PRESET_SCHEMA_VERSION = 1
PRESET_MAX_BYTES = 1_000_000  # 1 MB — real presets are a few KB
_VALID_PRESET_KEY = re.compile(r"^[A-Za-z][A-Za-z0-9_]*$")
# Streamlit-internal widget state that must not round-trip through presets.
_PRESET_KEY_BLOCKLIST_PREFIXES = ("FormSubmitter",)
_PRESET_KEY_BLOCKLIST_EXACT = {"preset_up"}


def _is_safe_preset_key(key) -> bool:
    """Return True if ``key`` looks like a legitimate preset field.

    Rejects non-strings, leading underscores (private/internal), Streamlit
    widget-state keys, and anything with characters outside
    ``[A-Za-z0-9_]``.
    """
    if not isinstance(key, str):
        return False
    if key.startswith("_"):
        return False
    if key in _PRESET_KEY_BLOCKLIST_EXACT:
        return False
    if any(key.startswith(p) for p in _PRESET_KEY_BLOCKLIST_PREFIXES):
        return False
    return bool(_VALID_PRESET_KEY.match(key))


def save_preset(session_state) -> str:
    """Save session state as JSON preset.

    Wraps the settings in an envelope with a schema version and timestamp
    so future loads can detect incompatibility instead of silently
    importing stale fields.

    Args:
        session_state: Streamlit session state

    Returns:
        JSON string of preset
    """
    payload = {
        k: v for k, v in session_state.items()
        if _is_safe_preset_key(k)
    }
    envelope = {
        "_preset_version": PRESET_SCHEMA_VERSION,
        "_saved_at": datetime.now(timezone.utc).isoformat(),
        "settings": payload,
    }
    return json.dumps(envelope, indent=2, default=str)


def load_preset(file, session_state) -> bool:
    """Load preset from JSON file.

    Validates schema version, caps total size, and accepts only keys that
    pass :func:`_is_safe_preset_key`. Unknown keys are surfaced as a
    warning; the load still succeeds for the valid subset so presets from
    earlier versions degrade gracefully rather than fail hard.

    Args:
        file: Uploaded JSON file
        session_state: Streamlit session state to update

    Returns:
        True if any valid settings were imported.
    """
    try:
        raw = file.read()
    except OSError as e:
        st.error(f"Failed to read preset file: {e}")
        return False

    if len(raw) > PRESET_MAX_BYTES:
        st.error(
            f"Preset file too large ({len(raw):,} bytes > "
            f"{PRESET_MAX_BYTES:,} byte limit)."
        )
        return False

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        st.error(f"Preset is not valid JSON: {e}")
        return False

    if not isinstance(data, dict):
        st.error("Preset must be a JSON object at the top level.")
        return False

    version = data.get("_preset_version")
    if version is None:
        # Older presets were a flat dict of settings. Accept them.
        settings = data
    elif version == PRESET_SCHEMA_VERSION:
        settings = data.get("settings", {})
        if not isinstance(settings, dict):
            st.error("Preset 'settings' field must be an object.")
            return False
    else:
        st.error(
            f"Preset version {version} is not compatible with this app "
            f"(expected {PRESET_SCHEMA_VERSION})."
        )
        return False

    accepted = 0
    rejected = []
    for key, value in settings.items():
        if _is_safe_preset_key(key):
            session_state[key] = value
            accepted += 1
        else:
            rejected.append(str(key))

    if rejected:
        preview = ", ".join(sorted(rejected)[:5])
        suffix = "..." if len(rejected) > 5 else ""
        st.warning(
            f"Ignored {len(rejected)} unrecognized preset key(s): "
            f"{preview}{suffix}"
        )
    return accepted > 0


def create_device_modifiers_excel(
    modifiers: Dict[Tuple[str, str], int],
    qmp_ids: list,
    filename_suffix: str = ""
) -> bytes:
    """Create Excel file for device modifiers upload.
    
    Args:
        modifiers: Dict of (qmp_id, device) -> bid percentage
        qmp_ids: List of QMP IDs to include
        filename_suffix: Optional suffix for filename
        
    Returns:
        Excel file bytes
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "source-modifiers"
    
    # Headers
    ws.cell(row=1, column=1, value="QMP ID")
    ws.cell(row=1, column=2, value="Mobile modifier")
    ws.cell(row=1, column=3, value="Desktop modifier")
    ws.cell(row=1, column=4, value="Tablet modifier")
    
    # Build rows
    row_num = 2
    for qmp_id in qmp_ids:
        ws.cell(row=row_num, column=1, value=qmp_id)
        
        # Get device modifiers (None = blank = 100%)
        mobile = modifiers.get((qmp_id, "mobile"))
        desktop = modifiers.get((qmp_id, "desktop"))
        tablet = modifiers.get((qmp_id, "tablet"))
        
        ws.cell(row=row_num, column=2, value=None if mobile is None else int(mobile))
        ws.cell(row=row_num, column=3, value=None if desktop is None else int(desktop))
        ws.cell(row=row_num, column=4, value=None if tablet is None else int(tablet))
        
        row_num += 1
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_existing_modifiers(
    file,
    required_columns: set = None
) -> Optional[Dict[Tuple[str, str], int]]:
    """Parse existing device modifiers from Excel file.
    
    Args:
        file: Uploaded Excel file
        required_columns: Set of required column names
        
    Returns:
        Dict of (canonical_qmp_id, device) -> modifier percentage, or None if invalid
    """
    if required_columns is None:
        required_columns = {
            "QMP ID",
            "Mobile modifier",
            "Desktop modifier",
            "Tablet modifier"
        }
    
    try:
        from data_utils import canonical_id
        
        # Read all sheets
        sheets = read_excel_file(file, sheet_name=None)
        
        # Find sheet with required columns
        target_df = None
        for sheet_df in sheets.values():
            if required_columns.issubset(set(sheet_df.columns)):
                target_df = sheet_df.copy()
                break
        
        if target_df is None:
            st.error(
                f"Excel file must contain sheet with columns: {', '.join(required_columns)}"
            )
            return None
        
        # Extract and parse
        df = target_df[list(required_columns)].copy()
        df["QMP_canonical"] = df["QMP ID"].map(canonical_id)
        
        # Coerce percentages (blank = 100)
        def coerce_percentage(col: pd.Series) -> pd.Series:
            numeric = pd.to_numeric(col, errors="coerce")
            numeric = numeric.fillna(100)
            return numeric.round().astype(int)
        
        df["Mobile modifier"] = coerce_percentage(df["Mobile modifier"])
        df["Desktop modifier"] = coerce_percentage(df["Desktop modifier"])
        df["Tablet modifier"] = coerce_percentage(df["Tablet modifier"])
        
        # Build mapping
        modifier_map = {}
        for _, row in df.iterrows():
            qmp_can = row["QMP_canonical"]
            if not qmp_can:
                continue
            
            modifier_map[(qmp_can, "mobile")] = int(row["Mobile modifier"])
            modifier_map[(qmp_can, "desktop")] = int(row["Desktop modifier"])
            modifier_map[(qmp_can, "tablet")] = int(row["Tablet modifier"])
        
        return modifier_map
        
    except Exception as e:
        st.error(f"Error parsing existing modifiers: {e}")
        return None
